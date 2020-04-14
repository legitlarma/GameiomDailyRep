from openpyxl import Workbook, load_workbook
import os
import pandas as pd
import numpy as np
import datetime
import xlrd
import time
def nextFreeRow(doc):                                                          
    count = 0                                                         
    filo = doc                 
    wb=xlrd.open_workbook(filo)                                       
    sheet=wb.sheet_by_index(0)                                        
    return (sheet.nrows)


def moveData(date, fileName, DLPath):

    time_to_wait = 10
    time_counter = 0
    while not os.path.exists(DLPath):
        time.sleep(1)
        time_counter += 1
    if time_counter > time_to_wait:
        print("Error - File Not Downloaded")
        return 0

    dest_filename = fileName
    b = False
    if os.path.isfile(dest_filename):
        wb = load_workbook(filename = dest_filename)
        b = True
    else:
        wb = Workbook()
    
    row = nextFreeRow(dest_filename)
    ws = wb.active

    ws.cell(row = row+1, column = 1).value = date
    tempDF = pd.read_csv(DLPath)
    tempDF = tempDF.replace(np.nan, '', regex=True)
    for i in range(len(tempDF)):
        row = row + 1
        tup = (())
        for j in range(len(tempDF.columns)):
            ws.cell(row = row+1, column = j+1).value = tempDF.iloc[i,j]

    wb.save(filename = fileName)

    os.remove(DLPath)

